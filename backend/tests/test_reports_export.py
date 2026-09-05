"""Excel export tests (Phase 9, T014-T015): workbook bytes, row fidelity
vs the JSON page, permission gates, empty result, bilingual headers, and
masking preserved inside the workbook."""

import io
import os
import uuid

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine, select
from sqlalchemy.orm import sessionmaker

from app.core.config import get_settings
from app.core.database import Base, dispose_engine, init_engine
from app.main import create_app
from app.modules.user.models import User
from app.seeds.seed_dev import run_seed

_TEST_DATABASE_URL = os.environ.get("DATABASE_URL")

requires_db = pytest.mark.skipif(
    not _TEST_DATABASE_URL or _TEST_DATABASE_URL.startswith("sqlite"),
    reason="export tests require real PostgreSQL",
)


@pytest.fixture()
def pg():
    engine = create_engine(_TEST_DATABASE_URL)  # type: ignore[arg-type]
    Base.metadata.create_all(engine)
    dispose_engine()
    init_engine(get_settings())
    factory = sessionmaker(bind=engine)
    with factory() as session:
        run_seed(session, prod=False)
    app = create_app()
    with TestClient(app) as test_client:
        yield test_client, factory
    dispose_engine()
    engine.dispose()


def _admin_token(client: TestClient) -> str:
    settings = get_settings()
    response = client.post(
        "/api/v1/auth/login",
        json={
            "email": settings.INITIAL_ADMIN_EMAIL,
            "password": settings.INITIAL_ADMIN_PASSWORD,
        },
    )
    assert response.status_code == 200, response.text
    return response.json()["access_token"]


def _bearer(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


def _setup_stock(client: TestClient, token: str) -> str:  # type: ignore[no-untyped-def]
    unique = uuid.uuid4().hex[:8]
    headers = _bearer(token)
    workplaces = client.get("/api/v1/org/workplaces?page_size=50", headers=headers)
    cp1 = next(w for w in workplaces.json()["items"] if w["code"] == "CP1")
    warehouse = client.post(
        "/api/v1/warehouse/warehouses",
        json={
            "workplace_id": cp1["id"],
            "code": f"WH-EXP-{unique}",
            "name": "Export WH",
            "name_fa": "انبار خروجی",
        },
        headers=headers,
    )
    shelf = client.post(
        f"/api/v1/warehouse/warehouses/{warehouse.json()['id']}/shelves",
        json={"code": "E-01"},
        headers=headers,
    )
    item = client.post(
        "/api/v1/warehouse/items",
        json={
            "name": f"Export item {unique}",
            "name_fa": f"کالای خروجی {unique}",
            "unit": "ad",
            "min_quantity": "0",
        },
        headers=headers,
    )
    receive = client.post(
        "/api/v1/warehouse/placements/receive",
        json={
            "item_id": item.json()["id"],
            "shelf_id": shelf.json()["id"],
            "quantity": "7",
            "reason": "export test",
        },
        headers=headers,
    )
    assert receive.status_code == 200, receive.text
    return item.json()["id"]


def _read_sheet(content: bytes) -> list[list]:  # type: ignore[type-arg]
    workbook = load_workbook(io.BytesIO(content), read_only=True)
    sheet = workbook.active
    rows = [[cell.value for cell in row] for row in sheet.iter_rows()]
    workbook.close()
    return rows


@requires_db
def test_export_inventory_matches_json_page(pg):  # type: ignore[no-untyped-def]
    client, _factory = pg
    token = _admin_token(client)
    item_id = _setup_stock(client, token)

    json_page = client.get(
        "/api/v1/reports/inventory?page_size=100", headers=_bearer(token)
    ).json()
    exported = client.get(
        "/api/v1/reports/export/excel?report=inventory&page_size=100",
        headers=_bearer(token),
    )
    assert exported.status_code == 200, exported.text
    assert exported.content[:2] == b"PK"  # xlsx zip magic bytes
    assert "spreadsheetml" in exported.headers["content-type"]
    assert "inventory-report-" in exported.headers["content-disposition"]
    assert exported.headers["content-disposition"].endswith('.xlsx"')

    rows = _read_sheet(exported.content)
    assert rows[0][0] == "item"  # en headers
    data = list(rows[1:])
    assert len(data) == json_page["total"]
    names = [row[0] for row in data]
    mine = next(
        row for row in json_page["items"] if row["item_id"] == item_id
    )
    assert mine["item_name"] in names


@requires_db
def test_export_empty_page_is_headers_only(pg):  # type: ignore[no-untyped-def]
    client, _factory = pg
    token = _admin_token(client)
    far_past = "2000-01-01T00:00:00Z"
    exported = client.get(
        f"/api/v1/reports/export/excel?report=requests&date_from={far_past}"
        f"&date_to=2000-01-02T00:00:00Z",
        headers=_bearer(token),
    )
    assert exported.status_code == 200
    rows = _read_sheet(exported.content)
    assert len(rows) == 1  # header row only
    assert rows[0][0] == "id"


@requires_db
def test_export_fa_locale_persian_headers_and_rtl(pg):  # type: ignore[no-untyped-def]
    client, _factory = pg
    token = _admin_token(client)
    _setup_stock(client, token)
    exported = client.get(
        "/api/v1/reports/export/excel?report=inventory&locale=fa&page_size=100",
        headers=_bearer(token),
    )
    assert exported.status_code == 200
    # RFC 5987: ASCII fallback in `filename`, UTF-8 Persian in `filename*`.
    disposition = exported.headers["content-disposition"]
    assert "inventory-report-" in disposition
    assert "filename*=UTF-8''" in disposition
    workbook = load_workbook(io.BytesIO(exported.content), read_only=False)
    sheet = workbook.active
    assert sheet.sheet_view.rightToLeft is True
    first_row = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
    assert first_row[0] == "کالا"
    workbook.close()


@requires_db
def test_export_requires_permission(pg):  # type: ignore[no-untyped-def]
    client, _factory = pg
    unauthenticated = client.get("/api/v1/reports/export/excel?report=inventory")
    assert unauthenticated.status_code == 401

    # roleless user
    from app.core.security import hash_password

    tag = uuid.uuid4().hex[:6]
    with pg[1]() as session:
        session.add(
            User(
                email=f"export-{tag}@zarandsteel.ir",
                username=f"export{tag}",
                hashed_password=hash_password("export-password-1"),
            )
        )
        session.commit()
    login = client.post(
        "/api/v1/auth/login",
        json={"email": f"export-{tag}@zarandsteel.ir", "password": "export-password-1"},
    )
    denied = client.get(
        "/api/v1/reports/export/excel?report=inventory",
        headers=_bearer(login.json()["access_token"]),
    )
    assert denied.status_code == 403

    admin_token = _admin_token(client)
    unknown = client.get(
        "/api/v1/reports/export/excel?report=bogus", headers=_bearer(admin_token)
    )
    assert unknown.status_code == 422
    assert unknown.json()["code"] == "VALIDATION_ERROR"


@requires_db
def test_export_audit_masking_preserved_in_workbook(pg):  # type: ignore[no-untyped-def]
    client, factory = pg
    token = _admin_token(client)
    headers = _bearer(token)
    unique = uuid.uuid4().hex[:8]
    ni = str(uuid.uuid4().int)[:10]

    workplaces = client.get("/api/v1/org/workplaces?page_size=50", headers=headers)
    cp1 = next(w for w in workplaces.json()["items"] if w["code"] == "CP1")
    employee = client.post(
        "/api/v1/employees",
        json={
            "national_id": ni,
            "personnel_code": f"EX-{unique}",
            "first_name": "Export",
            "last_name": "Mask",
            "workplace_id": cp1["id"],
            "user": {
                "email": f"export-mask-{unique}@zarandsteel.ir",
                "username": f"exportmask{unique}",
                "password": "export-mask-password-1",
            },
        },
        headers=headers,
    )
    assert employee.status_code == 201, employee.text

    # Admin (read_full) export contains masked snapshot text but never the
    # raw national id.
    exported = client.get(
        "/api/v1/reports/export/excel?report=audit&page_size=100",
        headers=headers,
    )
    assert exported.status_code == 200
    raw = exported.content.decode("utf-8", errors="ignore")
    assert ni not in raw  # the raw identifier never appears anywhere

    # Masked auditor (audit:log:read only, no read_full): snapshot cells empty.
    from app.core.security import hash_password
    from app.modules.user.models import Permission, Role, RolePermission, UserRole

    tag = uuid.uuid4().hex[:6]
    with factory() as session:
        user = User(
            email=f"exp-auditor-{tag}@zarandsteel.ir",
            username=f"expauditor{tag}",
            hashed_password=hash_password("exp-auditor-password-1"),
        )
        session.add(user)
        session.flush()
        role = Role(name=f"ExportAuditor-{tag}", description="")
        session.add(role)
        session.flush()
        for code in ["audit:log:read", "reports:export:excel"]:
            permission = session.scalar(
                select(Permission).where(Permission.code == code)
            )
            session.add(RolePermission(role_id=role.id, permission_id=permission.id))
        session.add(UserRole(user_id=user.id, role_id=role.id))
        session.commit()

    masked_login = client.post(
        "/api/v1/auth/login",
        json={
            "email": f"exp-auditor-{tag}@zarandsteel.ir",
            "password": "exp-auditor-password-1",
        },
    )
    masked_export = client.get(
        "/api/v1/reports/export/excel?report=audit&page_size=100",
        headers=_bearer(masked_login.json()["access_token"]),
    )
    assert masked_export.status_code == 200
    rows = _read_sheet(masked_export.content)
    header = rows[0]
    before_idx = header.index("before")
    after_idx = header.index("after")
    for row in rows[1:]:
        assert row[before_idx] in (None, "")
        assert row[after_idx] in (None, "")
